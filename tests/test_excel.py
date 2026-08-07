"""The workbook the user actually opens.

This is the only module that produces the deliverable, and pandas is coming out
of it in Etapa 8. These tests are the net for that: they describe the workbook,
not how it was built, so a rewrite that keeps the output passes and one that
quietly drops a sheet or a header does not.

Every test writes to a temporary directory. OUTPUT_DIR is read from the module
namespace at call time, so redirecting it there keeps the suite from leaving
files under the working directory.
"""

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from consulta_vacantes_mep.exports import excel as excel_module
from consulta_vacantes_mep.exports.excel import export_data_to_excel
from consulta_vacantes_mep.labels import APPOINTMENT_LABELS, VACANCY_LABELS
from consulta_vacantes_mep.models import Appointment, Vacancy
from consulta_vacantes_mep.settings import EXPORT


def _vacancy(
    number: str,
    specialty: str = "Francés",
    institution: str = "Liceo de Prueba",
) -> Vacancy:
    return Vacancy(
        number=number,
        regional_office="Dirección Regional de Prueba",
        position_class="Profesor de Enseñanza Media",
        specialty=specialty,
        institution=institution,
        lessons="10",
        starts_on="05/08/2026",
        ends_on="31/12/2026",
    )


def _appointment(vacancy_number: str) -> Appointment:
    """An appointment carrying an obviously invented national id.

    Real ones identify a specific person and have no business in a repository.
    """
    return Appointment(
        vacancy_number=vacancy_number,
        national_id="0-0000-0000",
        full_name="Persona De Prueba",
        institution="Liceo de Prueba",
        position_class="Profesor de Enseñanza Media",
        specialty="Francés",
        group="MT 4",
        position_number="0",
        starts_on="05/08/2026",
        ends_on="31/12/2026",
        status="Activo",
        eligibility_rating="0",
        roster_title="Nómina de prueba",
    )


PUBLISHED = [
    _vacancy("1531185"),
    _vacancy("1536996"),
    _vacancy("1538058", "Matemática"),
]


@pytest.fixture(autouse=True)
def output_dir(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    monkeypatch.setattr(excel_module, "OUTPUT_DIR", tmp_path)
    return tmp_path


def _row(sheet: Worksheet, number: int) -> list[object]:
    return [cell.value for cell in sheet[number]]


def _sheet(path: Path, name: str) -> Worksheet:
    return load_workbook(path)[name]


# ── Nothing to export ─────────────────────────────────────────────────────────
def test_no_vacancies_produces_no_workbook() -> None:
    assert export_data_to_excel([]) is None


def test_no_vacancies_writes_no_file(output_dir: Path) -> None:
    """An empty workbook is worse than no workbook: it looks like an answer."""
    export_data_to_excel([])

    assert list(output_dir.iterdir()) == []


# ── The file ──────────────────────────────────────────────────────────────────
def test_the_workbook_is_written_where_it_was_asked_for(output_dir: Path) -> None:
    path = export_data_to_excel(PUBLISHED)

    assert path is not None
    assert path.parent == output_dir
    assert path.exists()


def test_the_filename_carries_the_prefix_it_was_given() -> None:
    path = export_data_to_excel(PUBLISHED, filename_prefix="busqueda")

    assert path is not None
    assert path.name.startswith("busqueda_")
    assert path.suffix == ".xlsx"


def test_the_filename_defaults_to_the_spanish_prefix() -> None:
    """The user reads this one, so it stays in Spanish."""
    path = export_data_to_excel(PUBLISHED)

    assert path is not None
    assert path.name.startswith("vacantes_")


def test_the_filename_is_stamped_with_the_time() -> None:
    """Two runs an hour apart must not overwrite each other."""
    path = export_data_to_excel(PUBLISHED)

    assert path is not None
    stamp = path.stem.removeprefix("vacantes_")

    assert datetime.strptime(stamp, EXPORT.timestamp_format)  # noqa: DTZ007


# ── The sheets ────────────────────────────────────────────────────────────────
def test_both_sheets_are_always_present() -> None:
    path = export_data_to_excel(PUBLISHED)

    assert path is not None
    assert load_workbook(path).sheetnames == ["Vacantes", "Nombramientos"]


def test_the_vacancy_headers_are_the_spanish_labels() -> None:
    """labels.py is the single source of these, and the scrapers match table
    headers on the site against the same strings."""
    path = export_data_to_excel(PUBLISHED)

    assert path is not None
    assert _row(_sheet(path, "Vacantes"), 1) == list(VACANCY_LABELS.values())


def test_every_vacancy_gets_a_row() -> None:
    path = export_data_to_excel(PUBLISHED)

    assert path is not None
    assert _sheet(path, "Vacantes").max_row == len(PUBLISHED) + 1


def test_a_vacancy_row_holds_its_fields_in_label_order() -> None:
    path = export_data_to_excel([_vacancy("1531185")])

    assert path is not None
    assert _row(_sheet(path, "Vacantes"), 2) == [
        "1531185",
        "Dirección Regional de Prueba",
        "Profesor de Enseñanza Media",
        "Francés",
        "Liceo de Prueba",
        "10",
        "05/08/2026",
        "31/12/2026",
    ]


def test_the_appointment_headers_are_the_spanish_labels() -> None:
    path = export_data_to_excel(PUBLISHED, [_appointment("1531185")])

    assert path is not None
    assert _row(_sheet(path, "Nombramientos"), 1) == list(APPOINTMENT_LABELS.values())


def test_every_appointment_gets_a_row() -> None:
    """A vacancy can carry more than one, which is the point of cross-referencing."""
    appointments = [_appointment("1531185"), _appointment("1531185")]

    path = export_data_to_excel(PUBLISHED, appointments)

    assert path is not None
    assert _sheet(path, "Nombramientos").max_row == len(appointments) + 1


def test_an_appointments_sheet_with_nothing_in_it_has_no_headers() -> None:
    """Current behaviour, pinned rather than endorsed.

    A search where nobody was appointed yet produces a sheet with no column
    names at all, which reads as a broken export rather than an empty result.
    Worth revisiting when pandas comes out in Etapa 8; until then a rewrite
    should know it is changing this on purpose.
    """
    path = export_data_to_excel(PUBLISHED)

    assert path is not None
    assert _row(_sheet(path, "Nombramientos"), 1) == [None]


# ── The formatting ────────────────────────────────────────────────────────────
def test_the_header_row_is_filled_and_bold() -> None:
    path = export_data_to_excel(PUBLISHED)

    assert path is not None
    header = _sheet(path, "Vacantes")["A1"]

    # openpyxl reports colours with an alpha channel the settings do not carry.
    assert header.fill.start_color.rgb.endswith(EXPORT.header_fill_color)
    assert header.font.color.rgb.endswith(EXPORT.header_font_color)
    assert header.font.bold is True
    assert header.alignment.horizontal == "center"


def test_the_body_is_not_styled_like_the_header() -> None:
    path = export_data_to_excel(PUBLISHED)

    assert path is not None

    assert _sheet(path, "Vacantes")["A2"].font.bold is not True


def test_the_header_row_stays_visible_while_scrolling() -> None:
    """Fifty rows of vacancies are unreadable without it."""
    path = export_data_to_excel(PUBLISHED)

    assert path is not None
    assert _sheet(path, "Vacantes").freeze_panes == "A2"


def test_the_columns_can_be_filtered() -> None:
    path = export_data_to_excel(PUBLISHED)

    assert path is not None
    sheet = _sheet(path, "Vacantes")

    assert sheet.auto_filter.ref == sheet.dimensions


def test_a_long_value_does_not_widen_a_column_without_limit() -> None:
    """Institution names run long, and one of them should not push every other
    column off the screen."""
    path = export_data_to_excel([_vacancy("1531185", institution="Liceo " + "muy largo " * 20)])

    assert path is not None
    widths = _sheet(path, "Vacantes").column_dimensions

    assert widths["E"].width == EXPORT.max_column_width
