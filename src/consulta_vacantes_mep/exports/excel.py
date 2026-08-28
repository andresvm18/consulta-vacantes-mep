from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from consulta_vacantes_mep.labels import (
    APPOINTMENT_LABELS,
    appointment_to_row,
    vacancy_to_row,
)
from consulta_vacantes_mep.models import Appointment, Vacancy
from consulta_vacantes_mep.settings import EXPORT
from consulta_vacantes_mep.utils.logger import get_logger
from consulta_vacantes_mep.utils.paths import OUTPUT_DIR

logger = get_logger(__name__)


def format_worksheet(worksheet):
    header_fill = PatternFill(
        start_color=EXPORT.header_fill_color,
        end_color=EXPORT.header_fill_color,
        fill_type="solid",
    )

    header_font = Font(
        color=EXPORT.header_font_color,
        bold=True
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        adjusted_width = min(max_length + 2, EXPORT.max_column_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def export_data_to_excel(
    vacancies: list[Vacancy],
    appointments: list[Appointment] | None = None,
    filename_prefix: str = "vacantes",
) -> Path | None:
    if not vacancies:
        logger.warning("No vacancies to export; skipping workbook creation.")
        return None

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().astimezone().strftime(EXPORT.timestamp_format)
    filename = f"{filename_prefix}_{timestamp}.xlsx"
    file_path = OUTPUT_DIR / filename

    vacancies_df = pd.DataFrame([vacancy_to_row(v) for v in vacancies])
    appointments_df = pd.DataFrame(
        [appointment_to_row(a) for a in appointments or []],
        columns=list(APPOINTMENT_LABELS.values()),
    )

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        vacancies_df.to_excel(
            writer,
            sheet_name="Vacantes",
            index=False
        )

        appointments_df.to_excel(
            writer,
            sheet_name="Nombramientos",
            index=False
        )

        workbook = writer.book

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            format_worksheet(worksheet)

    return file_path
